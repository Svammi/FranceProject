from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import xlwings as xw
import math
from datetime import datetime

def mas_to_histogram(sheet,lenght):
    mas = []
    bin = []
    for i in range(2,lenght):
        if type(sheet.cell(row = i, column = 7).value) is not str:
            mas.append(sheet.cell(row = i, column = 7).value)
    print(mas)

    bin = np.arange(min(mas), max(mas) + 1, 0.05)
    return [mas,bin]

wb = load_workbook('./table.xlsx')
ws=wb['deceleration']
sheet_deceleration = wb.get_sheet_by_name('deceleration')

ws_acceleration = wb['acceleration']
sheet_acceleration = wb.get_sheet_by_name('acceleration')

deceleration = mas_to_histogram(sheet_deceleration, 21101)
acceleration = mas_to_histogram(sheet_acceleration, 24)

book = xw.Book()
sht=book.sheets[0]

title_graph = 'Histogram of the decelerations of all trains'
name_pictures = 'Plot'
fig, axs = plt.subplots(sharey=True, tight_layout=True, figsize=(10,3))
plt.xticks(deceleration[1], rotation='vertical')
weights = np.ones_like(deceleration[0])/float(len(deceleration[0])) #нормирование гистограмы
n, bins, patch = axs.hist(deceleration[0], bins=deceleration[1], weights = weights)
axs.set_xlim(0,2.6) # выборка значений Х для дальнейшего отсечения
axs.spines['right'].set_visible(False)
#axs.set( title=title_graph)
plt.xlabel("a\', m/s^2")
plt.ylabel("p")
sht.pictures.add(fig, name=name_pictures, update=True, top=sht.range('B'+str(1)).top)

title_graph = 'Histogram of the acceleartion of all trains'
name_pictures = 'Plot1'
fig, axs = plt.subplots(sharey=True, tight_layout=True, figsize=(10,3))
plt.xticks(acceleration[1], rotation='vertical')
weights = np.ones_like(acceleration[0])/float(len(acceleration[0])) #нормирование гистограмы
n, bins, patch = axs.hist(acceleration[0], bins=acceleration[1], weights = weights)
axs.set_xlim(0,2.6) # выборка значений Х для дальнейшего отсечения
axs.spines['right'].set_visible(False)
#axs.set( title=title_graph)
plt.xlabel("a\', m/s^2")
plt.ylabel("p")
sht.pictures.add(fig, name=name_pictures, update=True, top=sht.range('B'+str(30)).top)

