from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import xlwings as xw
import numpy as np
import copy
import matplotlib.ticker
import math
from datetime import datetime
from scipy.interpolate import interp1d


def calc_point_in_time(dict):
    dict_deceleration={}
    for i in dict.keys():
        dict_deceleration.update({i:[]})
    for i in dict.keys():
        time_start = dict[i][0][1]
        speed_start = dict[i][0][3]
        for j in range(0, len(dict[i])):
            second = datetime.strptime(dict[i][j][1], '%d/%m/%Y %H:%M:%S') - datetime.strptime(time_start, '%d/%m/%Y %H:%M:%S')
            dict[i][j][1]=second.seconds
            if 0<j<len(dict[i])-1:
                #deceleration_between = (speed_start - dict[i][j][3])/dict[i][j][1] #для торможения
                deceleration_between = (dict[i][j][3] - speed_start) / dict[i][j][1] #для ускорения
                dict_deceleration[i].append([dict[i][j][1],deceleration_between])

        dict_numpy = np.array(dict[i])
        dict[i] = dict_numpy.transpose()

        dict_numpy_deceleration = np.array(dict_deceleration[i])
        dict_deceleration[i] = dict_numpy_deceleration.transpose()

    return [dict, dict_deceleration]

def creat_dict_sequence(sheet,lenght):
    list_id_sequence_deceleration = []

    for k in range(2, lenght):  # range(2,184948):
        list_id_sequence_deceleration.append(sheet.cell(row=k, column=1).value)

    list_id_deceleration = list(set(list_id_sequence_deceleration))  # список id последовательностей

    dict_id_and_sequence_deceleration = {}
    for i in list_id_deceleration:
        dict_id_and_sequence_deceleration.update({i: []})

    for k in range(2, lenght):  # range(2,184948):
        dict_id_and_sequence_deceleration[sheet.cell(row=k, column=1).value].append(
            [sheet.cell(row=k, column=j).value for j in range(1, 7)])

    return dict_id_and_sequence_deceleration

def add_image_to_excel(book, dict_id_and_sequence_deceleration, dict_deceleration):
    k = 1
    for i in dict_id_and_sequence_deceleration.keys():
        if i==10 or i==11 or i==24237:
            x = np.float32(dict_id_and_sequence_deceleration[i][1])
            y_speed = np.float32(dict_id_and_sequence_deceleration[i][3])  # speed
            y_distance = np.float32(dict_id_and_sequence_deceleration[i][4])  # distance
            xi = np.linspace(x[0],x[-1], 50)  # промежуточных значение
            f_speed = interp1d(x, y_speed, kind='cubic')  # создание функции с помощью кубического интерполирования
            f_distance = interp1d(x, y_distance, kind='cubic')

            x_deceleration = dict_deceleration[i][0]  # chtp без крайних значений, т.к. для не вычисляется промежуточное значение торможения
            y_deceleration = dict_deceleration[i][1]
            xi_deceleration = np.linspace(x_deceleration[0], x_deceleration[-1], 100)
            f_deceleration = interp1d(x_deceleration, y_deceleration, kind='cubic')

            # отрисовка графиков зависимости расстояния от времени
            sht = book.sheets[0]
            title_graph = 'Graph of the distance traveled \n by train depending on the time \n for the sequence №' + str(i)
            name_pictures = 'Plot' + str(i)
            fig, ax = plt.subplots(figsize=(5,3))  # figsize=(3,3)
            plt.xticks(np.arange(0, max(x) + 1, 1), rotation='vertical')
            ax.plot(x, y_distance, 'bo', xi, f_distance(xi), '-', color="black")
            ax.set(xlabel='time (micro_s)', ylabel='distance (m)', title=title_graph)
            ax.grid()
            sht.pictures.add(fig, name=name_pictures, update=True, top=sht.range('B' + str((k - 1) * 21 + 1)).top)

            # отрисовка графиков зависимости скорости от времени
            sht = book.sheets[1]
            title_graph = 'The graph of the train speed \n depending on the time \n for the sequence №' + str(i)
            name_pictures = 'Plot' + str(i)
            fig, ax = plt.subplots(figsize=(5,3))  # figsize=(3,3)
            plt.xticks(np.arange(0, max(x) + 1, 1), rotation='vertical')
            ax.plot(x, y_speed, 'bo', xi, f_speed(xi), '-', color="black")
            ax.set(xlabel='time (micro_s)', ylabel='speed (V)', title=title_graph)
            ax.grid()
            sht.pictures.add(fig, name=name_pictures, update=True, top=sht.range('B' + str((k - 1) * 21 + 1)).top)

            # отрисовка графиков зависимости торможения от времени
            sht = book.sheets[2]
            title_graph = 'The graph of the train deceleartion \n depending on the time \n for the sequence №' + str(i)
            name_pictures = 'Plot' + str(i)
            fig, ax = plt.subplots(figsize=(5,3))  # figsize=(3,3)
            plt.xticks(np.arange(0, max(x_deceleration) + 1, 1), rotation='vertical')
            ax.plot(x_deceleration, y_deceleration, 'bo', xi_deceleration, f_deceleration(xi_deceleration), '-',
                    color='black')
            ax.set(xlabel='time(micro_s)', ylabel='deceleartion(D)', title=title_graph)
            ax.grid()
            sht.pictures.add(fig, name=name_pictures, update=True, top=sht.range('B' + str((k - 1) * 21 + 1)).top)
            k = k + 1


wb = load_workbook('./3_group_sequence.xlsx')
ws=wb['deceleration']
sheet_deceleration = wb.get_sheet_by_name('deceleration')
ws_acceleration = wb['acceleration']
sheet_acceleration = wb.get_sheet_by_name('acceleration')

dict_id_deceleration = creat_dict_sequence(sheet_deceleration,184948)
result = calc_point_in_time(dict_id_deceleration)
dict_id_deceleration = result[0]
deceleration = result[1]
book_deceleration = xw.Book()
add_image_to_excel(book_deceleration, dict_id_deceleration, deceleration)

dict_id_acceleration = creat_dict_sequence(sheet_acceleration,716)
result_acceleration = calc_point_in_time(dict_id_acceleration)
dict_id_acceleration = result_acceleration[0]
acceleration = result_acceleration[1]
book_acceleration = xw.Book()
add_image_to_excel(book_acceleration, dict_id_acceleration, acceleration)


