from openpyxl import load_workbook
from openpyxl import Workbook

'''загрузка документа'''
wb = load_workbook('./SCM_Train_Brake_DATABASE_2018_16_11.xlsx')
ws = wb['db']

sheet = wb.get_sheet_by_name('db')

#-------------------------------------------------создаю список id_sequense:value_sequence

list_id_sequence = []

for k in range(2,227639):
    list_id_sequence.append(sheet.cell(row=k, column=1).value)

list_id = list(set(list_id_sequence)) # список id последовательностей

dict_id_and_sequence={}
for i in list_id:
    dict_id_and_sequence.update({i:[]})

for k in range(2,227639):
    dict_id_and_sequence[sheet.cell(row=k,column=1).value].append([sheet.cell(row=k,column=j).value for j in range(1,7)])

#--------------------------------------------------------классификация  последовательностей-------------------------------------------------------#


id_sequense_where_have_none_value = []
for i in dict_id_and_sequence.keys():
    for sequence in dict_id_and_sequence[i]:
        if None in sequence:
            id_sequense_where_have_none_value.append(i)

list_id_with_none_value = list(set(id_sequense_where_have_none_value))
list_id_with_all_value = list(set(dict_id_and_sequence.keys()) - set(list_id_with_none_value))

wb_write = Workbook()
ws_deceleration = wb_write.create_sheet('deceleration')
ws_acceleration = wb_write.create_sheet('acceleration')
ws_null_value = wb_write.create_sheet('Have null value')

k=1
b=1
for i in list_id_with_all_value:
    if dict_id_and_sequence[i][0][3] - dict_id_and_sequence[i][-1][3] >=0:
        for j in range(0,len(dict_id_and_sequence[i])):
            ws_deceleration.cell(row=k+j, column=1).value = dict_id_and_sequence[i][j][0]
            ws_deceleration.cell(row=k+j, column=2).value = dict_id_and_sequence[i][j][1]
            ws_deceleration.cell(row=k+j, column=3).value = dict_id_and_sequence[i][j][2]
            ws_deceleration.cell(row=k+j, column=4).value = dict_id_and_sequence[i][j][3]
            ws_deceleration.cell(row=k+j, column=5).value = dict_id_and_sequence[i][j][4]
            ws_deceleration.cell(row=k+j, column=6).value = str(dict_id_and_sequence[i][j][5]*100)+'%'
        k=k+j+1
    else:
        for j in range(0,len(dict_id_and_sequence[i])):
            ws_acceleration.cell(row=b+j, column=1).value = dict_id_and_sequence[i][j][0]
            ws_acceleration.cell(row=b+j, column=2).value = dict_id_and_sequence[i][j][1]
            ws_acceleration.cell(row=b+j, column=3).value = dict_id_and_sequence[i][j][2]
            ws_acceleration.cell(row=b+j, column=4).value = dict_id_and_sequence[i][j][3]
            ws_acceleration.cell(row=b+j, column=5).value = dict_id_and_sequence[i][j][4]
            ws_acceleration.cell(row=b+j, column=6).value = str(dict_id_and_sequence[i][j][5]*100)+'%'
        b=b+j+1

k=1
for i in list_id_with_none_value:
    for j in range(0,len(dict_id_and_sequence[i])):
        ws_null_value.cell(row=k+j, column=1).value = dict_id_and_sequence[i][j][0]
        ws_null_value.cell(row=k+j, column=2).value = dict_id_and_sequence[i][j][1]
        ws_null_value.cell(row=k+j, column=3).value = dict_id_and_sequence[i][j][2]
        ws_null_value.cell(row=k+j, column=4).value = dict_id_and_sequence[i][j][3]
        ws_null_value.cell(row=k+j, column=5).value = dict_id_and_sequence[i][j][4]
        ws_null_value.cell(row=k+j, column=6).value = dict_id_and_sequence[i][j][5]
    k=k+j+1

wb_write.save('3_group_sequence.xlsx')

